import os
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


class VerificarVelocidade:
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_experimental_option("detach", True)
        servico = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=servico, options=chrome_options)

    def Inicializar(self):
        self.criar_cabecalho_planilha()
        self.acessar_site()

    def acessar_site(self):
        self.driver.get('https://fast.com/pt/')
        self.obter_informacoes()

    def obter_informacoes(self):
        # Tag xpath Download:
        # Tag xpath Upload:
        Download = self.driver.find_element(By.XPATH, '//*[@id="speed-value"]')
        time.sleep(20)
        element = self.driver.find_element(By.XPATH, '//*[@id="show-more-details-link"]')
        element.click()
        time.sleep(20)
        Upload = self.driver.find_element(By.XPATH, '//*[@id="upload-value"]')
        self.salvar_informacoes(Download, Upload)

    def criar_cabecalho_planilha(self):
        self.planilha = openpyxl.Workbook()
        self.planilha.create_sheet('Velocidade')
        self.planilha_internet = self.planilha['Velocidade']
        self.planilha_internet.cell(row=1, column=1, value='Download')
        self.planilha_internet.cell(row=1, column=2, value='Upload')

    def salvar_informacoes(self, Download, Upload):
        dados = [
            Download.text,
            Upload.text,
        ]
        self.planilha_internet.append(dados)
        self.driver.quit()
        self.planilha.save('Dados-velocidade.xlsx')
        print("\nVerificação Terminada !!!\n")